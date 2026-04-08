"""
气泡标注后端服务
支持任意 OpenAI 兼容 API 或 Anthropic API
支持 CAD 文件格式转换：DXF / DWG / SVG → PNG
"""
import base64
import io
import json
import os
import re
import tempfile
from flask import Flask, request, jsonify
from flask_cors import CORS

app = Flask(__name__)
CORS(app)

ANALYZE_PROMPT = """你是一个工程图纸分析专家。请仔细分析这张工程图纸，识别所有尺寸标注、公差、直径、角度、表面粗糙度等信息。

请以 JSON 格式返回，格式如下：
{
  "annotations": [
    {
      "value": "Ø12H7",
      "upper_tol": "+0.018",
      "lower_tol": "0",
      "type": "直径",
      "x_pct": 0.35,
      "y_pct": 0.42
    }
  ]
}

字段说明：
- value: 标注的主要内容，如 Ø12、30、R5、45°、Ra1.6 等
- upper_tol: 上偏差，如 "+0.018"，没有则为空字符串
- lower_tol: 下偏差，如 "-0.007"，没有则为空字符串
- type: 类型，只能是以下之一：直径、半径、尺寸、角度、粗糙度、其他
- x_pct: 标注在图纸中水平位置（0.0=最左，1.0=最右）
- y_pct: 标注在图纸中垂直位置（0.0=最上，1.0=最下）

要求：
1. 识别图纸中所有标注，不遗漏
2. 若标注含 ±，拆分为 upper_tol="+值"，lower_tol="-值"
3. 坐标指向标注文字位置
4. 只返回 JSON，不要有其他文字
"""

TYPE_COLORS = {
    "直径":  "#3498db", "半径": "#9b59b6", "尺寸": "#e67e22",
    "角度":  "#2ecc71", "粗糙度": "#1abc9c", "其他": "#607d8b",
}


def call_openai_compatible(api_key, base_url, model, image_b64, media_type, prompt=None):
    """调用 OpenAI 兼容接口（包括 OpenAI / Azure / 第三方）"""
    from openai import OpenAI
    client = OpenAI(api_key=api_key, base_url=base_url)
    response = client.chat.completions.create(
        model=model,
        max_tokens=4096,
        messages=[{
            "role": "user",
            "content": [
                {
                    "type": "image_url",
                    "image_url": {"url": f"data:{media_type};base64,{image_b64}"},
                },
                {"type": "text", "text": prompt or ANALYZE_PROMPT},
            ],
        }],
    )
    return response.choices[0].message.content


def call_anthropic(api_key, base_url, model, image_b64, media_type, prompt=None):
    """调用 Anthropic API"""
    import anthropic
    kwargs = {"api_key": api_key}
    if base_url:
        kwargs["base_url"] = base_url
    client = anthropic.Anthropic(**kwargs)
    response = client.messages.create(
        model=model,
        max_tokens=4096,
        messages=[{
            "role": "user",
            "content": [
                {
                    "type": "image",
                    "source": {
                        "type": "base64",
                        "media_type": media_type,
                        "data": image_b64,
                    },
                },
                {"type": "text", "text": prompt or ANALYZE_PROMPT},
            ],
        }],
    )
    return response.content[0].text


@app.route("/api/analyze", methods=["POST"])
def analyze():
    data = request.json or {}

    # ── 从请求体读取配置（前端设置面板传入）──────────────────────────────────
    api_key  = data.get("api_key")  or os.environ.get("API_KEY") or os.environ.get("ANTHROPIC_API_KEY") or os.environ.get("OPENAI_API_KEY")
    base_url = data.get("base_url") or os.environ.get("API_BASE_URL", "")
    model    = data.get("model")    or os.environ.get("API_MODEL", "claude-opus-4-6")
    provider = data.get("provider") or os.environ.get("API_PROVIDER", "anthropic")  # anthropic | openai
    image_b64 = data.get("image", "")

    if not api_key:
        return jsonify({"error": "未提供 API Key，请在设置面板填写或设置环境变量 API_KEY"}), 400
    if not image_b64:
        return jsonify({"error": "缺少 image 字段"}), 400

    # ── 处理 data URL 前缀 ───────────────────────────────────────────────────
    media_type = "image/png"
    if "," in image_b64:
        header, image_b64 = image_b64.split(",", 1)
        m = re.search(r"data:([^;]+);", header)
        if m:
            media_type = m.group(1)

    # ── 调用模型 ──────────────────────────────────────────────────────────────
    try:
        if provider == "anthropic":
            raw = call_anthropic(api_key, base_url or None, model, image_b64, media_type)
        else:
            # openai 或任意兼容接口
            raw = call_openai_compatible(api_key, base_url or "https://api.openai.com/v1", model, image_b64, media_type)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    # ── 解析 JSON ──────────────────────────────────────────────────────────────
    raw = raw.strip()
    json_match = re.search(r'\{.*\}', raw, re.DOTALL)
    if not json_match:
        return jsonify({"error": "模型未返回有效 JSON", "raw": raw}), 500

    try:
        result = json.loads(json_match.group())
    except json.JSONDecodeError as e:
        return jsonify({"error": f"JSON 解析失败: {e}", "raw": raw}), 500

    annotations = result.get("annotations", [])
    for ann in annotations:
        ann["color"] = TYPE_COLORS.get(ann.get("type", "other"), "#607d8b")

    return jsonify({"annotations": annotations, "count": len(annotations)})


@app.route("/api/convert", methods=["POST"])
def convert_cad():
    """将 CAD 文件（DXF/DWG/SVG）转换为 PNG base64"""
    if "file" not in request.files:
        return jsonify({"error": "缺少 file 字段"}), 400

    f = request.files["file"]
    filename = f.filename.lower()
    data = f.read()

    try:
        if filename.endswith(".svg"):
            png_bytes = _svg_to_png(data)

        elif filename.endswith(".dxf"):
            png_bytes = _dxf_to_png(data)

        elif filename.endswith(".dwg"):
            # ezdxf 0.19+ 支持读取 DWG（部分版本）
            png_bytes = _dwg_to_png(data, filename)

        else:
            return jsonify({"error": f"不支持的格式：{filename}"}), 400

        b64 = base64.b64encode(png_bytes).decode()
        return jsonify({"image": f"data:image/png;base64,{b64}"})

    except Exception as e:
        return jsonify({"error": str(e)}), 500


def _svg_to_png(data: bytes) -> bytes:
    try:
        import cairosvg
        return cairosvg.svg2png(bytestring=data, scale=2)
    except ImportError:
        # fallback：用 Pillow + cairosvg 不可用时尝试 Inkscape CLI
        import subprocess, shutil
        if shutil.which("inkscape"):
            with tempfile.NamedTemporaryFile(suffix=".svg", delete=False) as sf:
                sf.write(data); svg_path = sf.name
            out_path = svg_path.replace(".svg", ".png")
            subprocess.run(["inkscape", svg_path, "--export-filename", out_path,
                            "--export-dpi", "150"], check=True, capture_output=True)
            with open(out_path, "rb") as pf:
                return pf.read()
        raise RuntimeError("SVG 转换需要 cairosvg 或 inkscape，请运行：pip install cairosvg")


def _dxf_to_png(data: bytes) -> bytes:
    import ezdxf
    from ezdxf.addons.drawing import RenderContext, Frontend
    from ezdxf.addons.drawing.matplotlib import MatplotlibBackend
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    with tempfile.NamedTemporaryFile(suffix=".dxf", delete=False) as tf:
        tf.write(data); dxf_path = tf.name

    doc = ezdxf.readfile(dxf_path)
    msp = doc.modelspace()

    fig = plt.figure(figsize=(16, 12), dpi=150, facecolor="white")
    ax = fig.add_axes([0, 0, 1, 1])
    ctx = RenderContext(doc)
    out = MatplotlibBackend(ax)
    Frontend(ctx, out).draw_layout(msp, finalize=True)

    buf = io.BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight", facecolor="white")
    plt.close(fig)
    os.unlink(dxf_path)
    return buf.getvalue()


def _dwg_to_png(data: bytes, filename: str) -> bytes:
    import ezdxf
    # ezdxf 尝试直接读取 DWG（需要 ezdxf >= 0.19 且文件版本 <= R2018）
    with tempfile.NamedTemporaryFile(suffix=".dwg", delete=False) as tf:
        tf.write(data); dwg_path = tf.name
    try:
        doc = ezdxf.readfile(dwg_path)
    except Exception:
        os.unlink(dwg_path)
        raise RuntimeError("DWG 读取失败，请将文件另存为 DXF 格式后重试（AutoCAD: 文件→另存为→DXF）")

    from ezdxf.addons.drawing import RenderContext, Frontend
    from ezdxf.addons.drawing.matplotlib import MatplotlibBackend
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    msp = doc.modelspace()
    fig = plt.figure(figsize=(16, 12), dpi=150, facecolor="white")
    ax = fig.add_axes([0, 0, 1, 1])
    ctx = RenderContext(doc)
    out = MatplotlibBackend(ax)
    Frontend(ctx, out).draw_layout(msp, finalize=True)

    buf = io.BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight", facecolor="white")
    plt.close(fig)
    os.unlink(dwg_path)
    return buf.getvalue()


@app.route("/api/test", methods=["POST"])
def test_connection():
    data = request.json or {}
    api_key  = data.get("api_key", "").strip()
    base_url = data.get("base_url", "").strip()
    model    = data.get("model", "").strip()
    provider = data.get("provider", "openai")

    if not api_key:
        return jsonify({"ok": False, "error": "API Key 为空"}), 400

    try:
        if provider == "anthropic":
            import anthropic
            kwargs = {"api_key": api_key}
            if base_url:
                kwargs["base_url"] = base_url
            client = anthropic.Anthropic(**kwargs)
            resp = client.messages.create(
                model=model or "claude-opus-4-6",
                max_tokens=16,
                messages=[{"role": "user", "content": "hi"}],
            )
            reply = resp.content[0].text.strip()
        else:
            from openai import OpenAI
            client = OpenAI(
                api_key=api_key,
                base_url=base_url or "https://api.openai.com/v1",
            )
            resp = client.chat.completions.create(
                model=model or "gpt-4o",
                max_tokens=16,
                messages=[{"role": "user", "content": "hi"}],
            )
            reply = resp.choices[0].message.content.strip()

        return jsonify({"ok": True, "reply": reply})

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 200


@app.route("/api/extract-meta", methods=["POST"])
def extract_meta():
    """用 AI 从图片中提取图纸元信息（名称/图号/材质/数量/样品/批量）"""
    data       = request.get_json()
    image_b64  = data.get("image_b64", "")
    media_type = data.get("media_type", "image/jpeg")
    provider   = data.get("provider", "openai")
    base_url   = data.get("base_url", "")
    api_key    = data.get("api_key", "")
    model      = data.get("model", "")

    prompt = (
        "请从这张工程图纸或检验记录中提取以下信息，以 JSON 格式返回，"
        "字段名使用英文，值用图中原文：\n"
        "- name: 零件/产品名称\n"
        "- drawing: 图号或图纸编号\n"
        "- material: 材质\n"
        "- quantity: 数量（纯数字）\n"
        "- sample: 样品编号（若无则空字符串）\n"
        "- batch: 批量/批次号（若无则空字符串）\n\n"
        "只返回 JSON 对象，不要任何说明文字。"
    )

    try:
        if provider == "anthropic":
            raw = call_anthropic(api_key, base_url, model, image_b64, media_type, prompt)
        else:
            raw = call_openai_compatible(api_key, base_url, model, image_b64, media_type, prompt)

        # 从响应中提取 JSON
        match = re.search(r'\{.*\}', raw, re.DOTALL)
        if match:
            meta = json.loads(match.group())
        else:
            meta = {}
        return jsonify(meta)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/export", methods=["POST"])
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.page import PageMargins
    from flask import send_file

    data        = request.get_json()
    annotations = data.get("annotations", [])
    meta        = data.get("meta", {})
    name        = meta.get("name", "")
    drawing     = meta.get("drawing", "")
    material    = meta.get("material", "")
    quantity    = meta.get("quantity", "")
    sample      = meta.get("sample", "")
    batch       = meta.get("batch", "")
    date        = meta.get("date", "")
    inspector   = meta.get("inspector", "")
    reviewer    = meta.get("reviewer",  "")

    wb = Workbook()
    ws = wb.active
    ws.title = "日常检验记录"

    # ── 基础样式 ──
    THIN   = Side(style="thin")
    MEDIUM = Side(style="medium")
    T_BDR  = Border(left=THIN,   right=THIN,   top=THIN,   bottom=THIN)
    M_BDR  = Border(left=MEDIUM, right=MEDIUM, top=MEDIUM, bottom=MEDIUM)
    HDR_FILL = PatternFill("solid", fgColor="D9D9D9")
    SONG = "宋体"

    def sc(row, col, value="", bold=False, size=11,
           ha="center", va="center", fill=None):
        """Write and style a cell (border applied later in bulk)."""
        c = ws.cell(row=row, column=col, value=value)
        c.font      = Font(name=SONG, bold=bold, size=size)
        c.alignment = Alignment(horizontal=ha, vertical=va, wrap_text=False)
        if fill:
            c.fill = fill
        return c

    def fill_border(r1, c1, r2, c2, bdr=None):
        """Apply border to every cell in [r1:r2, c1:c2].
        Uses thin by default. This ensures merged-cell edges are correct."""
        b = bdr or T_BDR
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                ws.cell(r, c).border = b

    def merge(addr, value="", bold=False, size=11,
              ha="center", va="center", fill=None):
        """Merge a range, write value to anchor, apply borders to all cells."""
        ws.merge_cells(addr)
        # parse addr like "B3:E3"
        from openpyxl.utils.cell import range_boundaries
        min_c, min_r, max_c, max_r = range_boundaries(addr)
        c = sc(min_r, min_c, value, bold=bold, size=size, ha=ha, va=va, fill=fill)
        fill_border(min_r, min_c, max_r, max_c)
        return c

    # ── 列宽（10列）──
    # A=序号 B=尺寸 C=公差 D-H=实测1-5 I=OK J=NO
    widths = [7, 14, 11, 10, 10, 10, 10, 10, 7, 7]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ━━ R1: 大标题 ━━
    ws.row_dimensions[1].height = 36
    merge("A1:J1", "林海日常检验记录", bold=True, size=18)
    ws.cell(1, 1).border = Border()   # 标题行不要边框

    # ━━ R2: LHJQ / 样品 / 批量 / 日期（无边框，辅助信息行）━━
    ws.row_dimensions[2].height = 18
    sc(2, 1, "LHJQ", bold=True, size=10, ha="left")
    ws.merge_cells("B2:D2");  ws.cell(2, 2).value = f"样品: {sample}"
    ws.cell(2, 2).font = Font(name=SONG, size=10)
    ws.cell(2, 2).alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("E2:G2");  ws.cell(2, 5).value = f"批量: {batch}"
    ws.cell(2, 5).font = Font(name=SONG, size=10)
    ws.cell(2, 5).alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("H2:J2");  ws.cell(2, 8).value = f"日期: {date}"
    ws.cell(2, 8).font = Font(name=SONG, size=10)
    ws.cell(2, 8).alignment = Alignment(horizontal="left", vertical="center")

    # ━━ R3: 名称 / 图号 ━━
    ws.row_dimensions[3].height = 24
    sc(3, 1, "名称", bold=True, size=11, fill=HDR_FILL)
    fill_border(3, 1, 3, 1)
    merge("B3:E3", name, size=11)
    sc(3, 6, "图号", bold=True, size=11, fill=HDR_FILL)
    fill_border(3, 6, 3, 6)
    merge("G3:J3", drawing, size=11)

    # ━━ R4: 材质 / 数量 ━━
    ws.row_dimensions[4].height = 24
    sc(4, 1, "材质", bold=True, size=11, fill=HDR_FILL)
    fill_border(4, 1, 4, 1)
    merge("B4:E4", material, size=11)
    sc(4, 6, "数量", bold=True, size=11, fill=HDR_FILL)
    fill_border(4, 6, 4, 6)
    merge("G4:I4", quantity, size=11)
    sc(4, 10, "件", size=11)
    fill_border(4, 10, 4, 10)

    # ━━ R5-R6: 数据表头（双行合并）━━
    ws.row_dimensions[5].height = 22
    ws.row_dimensions[6].height = 18
    # 序号/尺寸/公差/OK/NO 跨两行
    for col, label in [(1,"序号"),(2,"尺寸"),(3,"公差"),(9,"OK"),(10,"NO")]:
        addr = f"{get_column_letter(col)}5:{get_column_letter(col)}6"
        merge(addr, label, bold=True, size=11, fill=HDR_FILL)
    # 实测尺寸 横向合并
    merge("D5:H5", "实测尺寸", bold=True, size=11, fill=HDR_FILL)
    # 实测子列 1-5
    for i, sub in enumerate(["1","2","3","4","5"], 4):
        sc(6, i, sub, bold=True, size=10, fill=HDR_FILL)
        fill_border(6, i, 6, i)

    # ━━ R7+: 数据行 ━━
    for idx, ann in enumerate(annotations):
        r = 7 + idx
        ws.row_dimensions[r].height = 20
        # 公差格式化
        u = str(ann.get("upper_tol", "")).strip()
        l = str(ann.get("lower_tol", "")).strip()
        if u and l:
            try:
                if abs(float(u)) == abs(float(l)):
                    tol = f"±{abs(float(u))}"
                else:
                    tol = f"{u}/{l}"
            except ValueError:
                tol = f"{u}/{l}"
        else:
            tol = u or l or ""
        vals = [ann.get("num",""), ann.get("value",""), tol,
                "", "", "", "", "", "", ""]
        for c, v in enumerate(vals, 1):
            sc(r, c, v, size=11)
        fill_border(r, 1, r, 10)

    # ━━ R(7+n): 表尾——检验员 / 审核 ━━
    footer_row = 7 + len(annotations)
    ws.row_dimensions[footer_row].height = 22
    # 检验员：A-E
    sc(footer_row, 1, "检验员：", bold=True, size=11, ha="right")
    fill_border(footer_row, 1, footer_row, 1)
    merge(f"B{footer_row}:E{footer_row}", inspector, size=11, ha="left")
    # 审核：F-J
    sc(footer_row, 6, "审  核：", bold=True, size=11, ha="right")
    fill_border(footer_row, 6, footer_row, 6)
    merge(f"G{footer_row}:J{footer_row}", reviewer, size=11, ha="left")

    # ━━ 外框加粗 ━━
    last_row = footer_row
    # 整体表格区域（R3-last_row）加粗外框
    for r in range(3, last_row + 1):
        for c in range(1, 11):
            cell = ws.cell(r, c)
            b = cell.border
            cell.border = Border(
                left   = MEDIUM if c == 1  else b.left,
                right  = MEDIUM if c == 10 else b.right,
                top    = MEDIUM if r == 3  else b.top,
                bottom = MEDIUM if r == last_row else b.bottom,
            )

    # ━━ 打印设置 ━━
    total_rows = last_row
    ws.print_area = f"A1:J{total_rows}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize   = 9      # A4
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.59, right=0.59, top=0.79, bottom=0.79,
                                   header=0.31, footer=0.31)
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"林海日常检验记录_{date}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )


@app.route("/health")
def health():
    return jsonify({"status": "ok"})


if __name__ == "__main__":
    print("🚀 服务启动：http://localhost:5001")
    print("   支持通过前端设置面板配置 API Provider / Key / Model")
    app.run(port=5001, debug=True)
